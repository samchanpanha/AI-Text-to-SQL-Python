import os
import uuid
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.agents.sql_executor import execute_sql
from app.reports.formatter import get_column_format
from config.app import Settings


settings = Settings()


def generate_excel(
    sql_query: str,
    sheet_name: str = "Sheet1",
    output_dir: str | None = None,
    filename_prefix: str = "report",
) -> str:
    """Execute SQL and generate a formatted Excel file. Returns file path."""
    result = execute_sql(sql_query)
    if not result["success"]:
        raise RuntimeError(f"SQL execution failed: {result['error']}")

    return _write_excel(
        columns=result["columns"],
        rows=[list(r.values()) for r in result["results"]],
        sheet_name=sheet_name,
        output_dir=output_dir or settings.REPORT_TEMP_DIR,
        filename_prefix=filename_prefix,
    )


def _write_excel(
    columns: list[str],
    rows: list[list],
    sheet_name: str = "Sheet1",
    output_dir: str = "/tmp/reports",
    filename_prefix: str = "report",
) -> str:
    """Write data to a formatted Excel file using write-only mode for large datasets."""
    os.makedirs(output_dir, exist_ok=True)
    file_id = uuid.uuid4().hex[:8]
    filename = f"{filename_prefix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file_id}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name[:31])

    # Detect column formats from header names
    col_formats = [get_column_format(col_name) for col_name in columns]

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    header_row = []
    for col_name in columns:
        cell_value = col_name
        header_row.append(cell_value)

    ws.append(header_row)

    # Data rows — streaming (write-only), handles large datasets
    chunk_count = 0
    for i, row in enumerate(rows):
        formatted_row = []
        for j, value in enumerate(row):
            fmt = col_formats[j] if j < len(col_formats) else {}
            formatted_row.append(_format_cell_value(value, fmt))
        ws.append(formatted_row)

        chunk_count += 1
        if chunk_count >= settings.REPORT_STREAM_CHUNK:
            chunk_count = 0

    wb.save(filepath)
    return filepath


def _format_cell_value(value, fmt: dict):
    """Apply formatting rules to a cell value."""
    if value is None:
        return ""

    number_format = fmt.get("number_format")
    force_text = fmt.get("force_text", False)

    if force_text:
        return str(value)

    if number_format and isinstance(value, (int, float)):
        # Return raw value; openpyxl number_format is applied via styles
        # In write-only mode, we use string formatting instead
        if number_format == "#,##0":
            return int(value) if isinstance(value, float) else value
        elif number_format == "#,##0.00":
            return round(float(value), 2)
        elif number_format == "#,##0.0000":
            return round(float(value), 4)
        return value

    return value
